from __future__ import annotations

import math
import os
import re
import threading
import time
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any, Iterable
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unknown extension is not supported")
warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported")

STATUS_ICON_PROGRESS = {
    "○": 0.0,
    "◔": 25.0,
    "◑": 50.0,
    "◕": 75.0,
    "●": 100.0,
}

STATUS_LABELS = {
    "complete": "Completed",
    "active": "Active",
    "not_started": "Not started",
    "at_risk": "At risk",
    "blocked": "Blocked",
    "no_update": "No update",
}

EXCLUDED_SHEETS = {
    "Master",
    "Combined Deck",
    "MoS End to End",
    "List Format",
    "Portfolio",
    "KEY RISKS",
}


def _clean_text(value: Any, *, collapse: bool = False) -> str:
    if value is None:
        return ""
    text = str(value)
    text = (
        text.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\u00a0", " ")
        .replace("\u200b", "")
        .replace("\ufeff", "")
        .replace("\t", " ")
    )
    text = re.sub(r"[ \f\v]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = text.strip()
    if collapse:
        text = re.sub(r"\s+", " ", text).strip()
    return text


def _is_year(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if float(value).is_integer() and 2000 <= int(value) <= 2100:
            return int(value)
    if isinstance(value, str):
        match = re.fullmatch(r"\s*(20\d{2})\s*", value)
        if match:
            return int(match.group(1))
    return None


def _parse_progress(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return None
        if 0 <= number <= 1:
            number *= 100
        return round(max(0.0, min(100.0, number)), 1)

    text = _clean_text(value, collapse=True)
    if text in STATUS_ICON_PROGRESS:
        return STATUS_ICON_PROGRESS[text]

    match = re.search(r"(-?\d+(?:\.\d+)?)\s*%", text)
    if match:
        return round(max(0.0, min(100.0, float(match.group(1)))), 1)

    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        number = float(text)
        if 0 <= number <= 1:
            number *= 100
        if 0 <= number <= 100:
            return round(number, 1)
    return None


def _format_timeline(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    return _clean_text(value, collapse=True)


def _parse_due_date(value: Any, year_hint: int) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None

    text = _clean_text(value, collapse=True)
    if not text:
        return None

    # ISO and common date strings.
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d %b %Y",
        "%d %B %Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    # Q1, Q3 2026, Q3/Q4 2026: use the latest stated quarter.
    quarter_matches = re.findall(r"Q([1-4])", text, flags=re.IGNORECASE)
    if quarter_matches:
        quarter = max(int(q) for q in quarter_matches)
        year_match = re.search(r"(20\d{2})", text)
        year = int(year_match.group(1)) if year_match else year_hint
        month_day = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}[quarter]
        return date(year, *month_day)

    # Aug'26, August 2026.
    month_match = re.search(
        r"\b(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
        r"\s*[' -]?(\d{2,4})?\b",
        text,
        flags=re.IGNORECASE,
    )
    if month_match:
        month_name = month_match.group(1)[:3].title()
        month = datetime.strptime(month_name, "%b").month
        raw_year = month_match.group(2)
        if raw_year:
            parsed_year = int(raw_year)
            year = 2000 + parsed_year if parsed_year < 100 else parsed_year
        else:
            year = year_hint
        # Month-end without adding a dateutil dependency.
        next_month = date(year + (month == 12), 1 if month == 12 else month + 1, 1)
        return date.fromordinal(next_month.toordinal() - 1)

    return None


def _status_from(progress: float | None, timeline_values: Iterable[Any], year_hint: int) -> str:
    if progress is None:
        return "no_update"
    if progress >= 99.5:
        return "complete"
    if progress <= 0.5:
        return "not_started"

    due_dates = [
        due
        for value in timeline_values
        if (due := _parse_due_date(value, year_hint)) is not None
    ]
    if due_dates and min(due_dates) < date.today():
        return "at_risk"
    return "active"


def _split_items(value: Any) -> list[str]:
    text = _clean_text(value)
    if not text:
        return []

    # Put every bullet or numbered item on a new line, even when the source cell
    # contains multiple markers without clean line breaks.
    text = re.sub(r"\s*([•▪◦])\s*", r"\n\1 ", text)
    text = re.sub(r"(?m)(?<!\w)(\d{1,2}[.)])\s+", r"\n\1 ", text)

    items: list[str] = []
    for raw_line in re.split(r"\n+", text):
        line = raw_line.strip()
        line = re.sub(r"^[•▪◦●\-–—]+\s*", "", line)
        line = re.sub(r"^\d{1,2}[.)]\s*", "", line)
        line = re.sub(r"\s+", " ", line).strip(" ;|/")
        if not line or not re.search(r"[A-Za-z0-9]", line):
            continue
        if line.lower() in {"na", "n/a", "none", "tbd"}:
            continue
        items.append(line)

    # Preserve order while removing exact duplicate lines.
    seen: set[str] = set()
    unique: list[str] = []
    for item in items:
        key = item.casefold()
        if key not in seen:
            unique.append(item)
            seen.add(key)
    return unique


def _average(values: Iterable[float | None]) -> float | None:
    cleaned = [float(v) for v in values if v is not None]
    return round(mean(cleaned), 1) if cleaned else None


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug[:70] or "item"


def _is_metric_block_header(ws, row: int) -> bool:
    a = _clean_text(ws.cell(row, 1).value, collapse=True).lower()
    b = _clean_text(ws.cell(row, 2).value, collapse=True).lower()
    c = _clean_text(ws.cell(row, 3).value, collapse=True).lower()
    if "metric" in a and ("function" in a or "interdepend" in a):
        return True
    if b in {"owner", "OWNER".lower()} and "baseline" in c:
        return True
    return False


def _merged_end_row(ws, row: int, col: int) -> int | None:
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row == row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return merged_range.max_row
    return None


def _dependency_groups(ws, header_row: int) -> list[tuple[int, str]]:
    groups: list[tuple[int, str]] = []
    for col in range(8, max(8, ws.max_column - 2)):
        function_name = _clean_text(ws.cell(header_row, col).value, collapse=True)
        timeline_header = _clean_text(ws.cell(header_row, col + 1).value, collapse=True).lower()
        tracker_header = _clean_text(ws.cell(header_row, col + 2).value, collapse=True).lower()
        if function_name and timeline_header in {"time line", "timeline"} and "progress" in tracker_header:
            groups.append((col, function_name))
    return groups


def _theme_rows(ws, start_row: int, end_row: int) -> list[tuple[int, int]]:
    starts: list[int] = []
    for row in range(start_row, end_row + 1):
        value = _clean_text(ws.cell(row, 1).value, collapse=True)
        if not value:
            continue
        if _is_year(value) is not None:
            continue
        lower = value.lower()
        if lower == "key themes" or ("metric" in lower and ("function" in lower or "interdepend" in lower)):
            continue
        starts.append(row)

    result: list[tuple[int, int]] = []
    for index, row in enumerate(starts):
        merged_end = _merged_end_row(ws, row, 1)
        if merged_end is not None:
            theme_end = min(merged_end, end_row)
        elif index + 1 < len(starts):
            theme_end = starts[index + 1] - 1
        else:
            theme_end = end_row
        result.append((row, theme_end))
    return result


def _row_status_records(ws, start_row: int, end_row: int, status_col: int, progress_col: int, timeline_col: int) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row in range(start_row, end_row + 1):
        status_value = ws.cell(row, status_col).value
        progress_value = ws.cell(row, progress_col).value
        timeline_value = ws.cell(row, timeline_col).value
        progress = _parse_progress(progress_value)
        if progress is None:
            progress = _parse_progress(status_value)
        if status_value in (None, "") and progress_value in (None, "") and timeline_value in (None, ""):
            continue
        records.append(
            {
                "row": row,
                "progress": progress,
                "raw_status": _clean_text(status_value, collapse=True),
                "timeline_raw": timeline_value,
                "timeline": _format_timeline(timeline_value),
            }
        )
    return records


def _combine_statuses(actions: list[dict[str, Any]]) -> str:
    if not actions:
        return "no_update"
    if all(act["status"] == "complete" for act in actions):
        return "complete"
    status_priority = ["blocked", "at_risk", "active", "not_started", "no_update"]
    for status in status_priority:
        if any(act["status"] == status for act in actions):
            return status
    return "no_update"


def _parse_theme(ws, function_name: str, year: int, start_row: int, end_row: int, dependency_groups: list[tuple[int, str]]) -> dict[str, Any]:
    theme_name = _clean_text(ws.cell(start_row, 1).value, collapse=True)

    action_texts: list[str] = []
    for row in range(start_row, end_row + 1):
        action_texts.extend(_split_items(ws.cell(row, 2).value))

    own_records = _row_status_records(ws, start_row, end_row, 5, 6, 7)
    own_progress_values = [record["progress"] for record in own_records]
    theme_progress = _average(own_progress_values)
    own_timelines_raw = [record["timeline_raw"] for record in own_records if record["timeline_raw"] not in (None, "")]
    theme_status = _status_from(theme_progress, own_timelines_raw, year)

    actions: list[dict[str, Any]] = []
    action_count = max(len(action_texts), len(own_records))
    for index in range(action_count):
        record = own_records[index] if index < len(own_records) else None
        text = action_texts[index] if index < len(action_texts) else f"Action item {index + 1}"
        progress = record["progress"] if record else None
        timeline_raw = record["timeline_raw"] if record else None
        status = _status_from(progress, [timeline_raw] if timeline_raw not in (None, "") else [], year)
        actions.append(
            {
                "id": f"{_slug(theme_name)}-action-{index + 1}",
                "description": text,
                "progress": progress,
                "status": status,
                "status_label": STATUS_LABELS[status],
                "timeline": record["timeline"] if record else "",
                "owner": "Not assigned",
            }
        )

    dependencies: list[dict[str, Any]] = []
    for group_col, dependency_function in dependency_groups:
        row_action_texts = []
        for row in range(start_row, end_row + 1):
            cell_val = ws.cell(row, group_col).value
            if cell_val not in (None, ""):
                row_action_texts.extend(_split_items(cell_val))

        row_trackers = []
        for row in range(start_row, end_row + 1):
            timeline_value = ws.cell(row, group_col + 1).value
            status_value = ws.cell(row, group_col + 2).value
            progress_value = ws.cell(row, group_col + 3).value
            
            is_empty = timeline_value in (None, "") and status_value in (None, "") and progress_value in (None, "")
            if not is_empty:
                row_trackers.append({
                    "timeline": timeline_value,
                    "status": status_value,
                    "progress": progress_value,
                })

        if not row_action_texts and not row_trackers:
            continue

        num_items = max(len(row_action_texts), len(row_trackers), 1)
        actions_list = []
        for i in range(num_items):
            desc = row_action_texts[i] if i < len(row_action_texts) else f"Action item {i + 1}"
            prog = None
            stat = "no_update"
            time_str = ""
            timeline_raw = None
            
            if i < len(row_trackers):
                tracker = row_trackers[i]
                timeline_raw = tracker["timeline"]
                time_str = _format_timeline(timeline_raw)
                prog = _parse_progress(tracker["progress"])
                if prog is None:
                    prog = _parse_progress(tracker["status"])
                stat = _status_from(prog, [timeline_raw] if timeline_raw not in (None, "") else [], year)

            actions_list.append({
                "description": desc,
                "progress": prog,
                "status": stat,
                "status_label": STATUS_LABELS[stat],
                "timeline": time_str,
                "owner": "Not assigned",
                "notes": "—",
            })

        dep_progress = _average([act["progress"] for act in actions_list if act["progress"] is not None])
        dep_status = _combine_statuses(actions_list)
        timeline_labels = []
        for act in actions_list:
            if act["timeline"] and act["timeline"] not in timeline_labels:
                timeline_labels.append(act["timeline"])

        dependencies.append(
            {
                "id": f"{_slug(theme_name)}-{_slug(dependency_function)}",
                "function": dependency_function,
                "actions": actions_list,
                "progress": dep_progress,
                "status": dep_status,
                "status_label": STATUS_LABELS[dep_status],
                "timeline": " · ".join(timeline_labels),
                "owner": "Not assigned",
                "source_column": get_column_letter(group_col),
            }
        )

    dependency_progress = _average(dep["progress"] for dep in dependencies)
    dependency_status = _status_from(
        dependency_progress,
        [dep["timeline"] for dep in dependencies if dep["timeline"]],
        year,
    )

    timelines = []
    for record in own_records:
        if record["timeline"] and record["timeline"] not in timelines:
            timelines.append(record["timeline"])

    return {
        "id": f"{_slug(function_name)}-{year}-{start_row}",
        "row": start_row,
        "name": theme_name,
        "progress": theme_progress,
        "status": theme_status,
        "status_label": STATUS_LABELS[theme_status],
        "timeline": " · ".join(timelines),
        "actions": actions,
        "dependencies": dependencies,
        "dependency_progress": dependency_progress,
        "dependency_status": dependency_status,
        "dependency_status_label": STATUS_LABELS[dependency_status],
        "risk_count": sum(1 for dep in dependencies if dep["status"] in {"at_risk", "blocked"}),
        "missing_update_count": sum(1 for dep in dependencies if dep["status"] == "no_update"),
    }


def _find_year_sections(ws) -> list[tuple[int, int]]:
    sections: list[tuple[int, int]] = []
    for row in range(1, ws.max_row):
        year = _is_year(ws.cell(row, 1).value)
        next_header = _clean_text(ws.cell(row + 1, 1).value, collapse=True).lower()
        if year and next_header == "key themes":
            sections.append((row, year))
    return sections


def _parse_function_sheet(ws) -> dict[str, Any] | None:
    year_rows = _find_year_sections(ws)
    if not year_rows:
        return None

    # Use the worksheet tab name so the dashboard filter exactly matches Excel.
    function_name = ws.title
    years: dict[str, Any] = {}

    for index, (year_row, year) in enumerate(year_rows):
        header_row = year_row + 1
        data_start = header_row + 2
        next_year_row = year_rows[index + 1][0] if index + 1 < len(year_rows) else ws.max_row + 1
        data_end = next_year_row - 1

        for row in range(data_start, data_end + 1):
            if _is_metric_block_header(ws, row):
                data_end = row - 1
                break

        groups = _dependency_groups(ws, header_row)
        themes = [
            _parse_theme(ws, function_name, year, start_row, end_row, groups)
            for start_row, end_row in _theme_rows(ws, data_start, data_end)
        ]

        dep_counters: dict[str, int] = {}
        sheet_short = "".join([w[0] for w in ws.title.split() if w.isalnum()]).upper()
        if not sheet_short:
            sheet_short = "DEP"

        for theme in themes:
            for dep in theme.get("dependencies", []):
                func = dep["function"]
                if func not in dep_counters:
                    dep_counters[func] = 0
                for action in dep.get("actions", []):
                    dep_counters[func] += 1
                    func_clean = re.sub(r'[^a-zA-Z0-9]', '', func).upper()
                    action["id"] = f"{sheet_short}-{func_clean}-{dep_counters[func]:02d}"

        summary = {
            "total": len(themes),
            "active": sum(1 for theme in themes if theme["status"] in {"active", "at_risk", "blocked"}),
            "not_started": sum(1 for theme in themes if theme["status"] == "not_started"),
            "completed": sum(1 for theme in themes if theme["status"] == "complete"),
            "no_update": sum(1 for theme in themes if theme["status"] == "no_update"),
            "at_risk": sum(1 for theme in themes if theme["status"] in {"at_risk", "blocked"}),
            "average_progress": _average(theme["progress"] for theme in themes),
            "dependency_count": sum(len(theme["dependencies"]) for theme in themes),
            "completed_dependencies": sum(
                1
                for theme in themes
                for dependency in theme["dependencies"]
                if dependency["status"] == "complete"
            ),
        }
        years[str(year)] = {"year": year, "summary": summary, "themes": themes}

    return {
        "name": function_name,
        "sheet_name": ws.title,
        "years": years,
    }


@dataclass
class WorkbookDashboardStore:
    workbook_path: Path
    poll_retry_count: int = 4
    poll_retry_delay: float = 0.35

    def __post_init__(self) -> None:
        self.workbook_path = Path(self.workbook_path)
        self._lock = threading.RLock()
        self._mtime_ns: int | None = None
        self._size: int | None = None
        self._snapshot: dict[str, Any] | None = None
        self._last_error: str | None = None

    def _version(self) -> tuple[int, int]:
        stat = self.workbook_path.stat()
        return stat.st_mtime_ns, stat.st_size

    def _load(self) -> dict[str, Any]:
        workbook = None
        last_exception: Exception | None = None
        for attempt in range(self.poll_retry_count):
            try:
                workbook = load_workbook(
                    self.workbook_path,
                    data_only=True,
                    read_only=False,
                    keep_links=False,
                )
                break
            except (BadZipFile, PermissionError, OSError) as exc:
                last_exception = exc
                if attempt + 1 < self.poll_retry_count:
                    time.sleep(self.poll_retry_delay)
        if workbook is None:
            raise RuntimeError(f"Could not read workbook after retrying: {last_exception}")

        functions: dict[str, Any] = {}
        for ws in workbook.worksheets:
            if ws.title in EXCLUDED_SHEETS:
                continue
            parsed = _parse_function_sheet(ws)
            if parsed:
                functions[parsed["name"]] = parsed

        workbook.close()

        ordered_function_names = list(functions.keys())
        preferred = "Service Transformation"
        default_function = preferred if preferred in functions else next(
            (name for name in ordered_function_names if any(year["summary"]["total"] for year in functions[name]["years"].values())),
            ordered_function_names[0] if ordered_function_names else "",
        )

        years_for_default = functions.get(default_function, {}).get("years", {})
        default_year = ""
        if years_for_default:
            populated_years = [key for key, value in years_for_default.items() if value["summary"]["total"] > 0]
            default_year = populated_years[0] if populated_years else next(iter(years_for_default))

        mtime_ns, size = self._version()
        modified = datetime.fromtimestamp(mtime_ns / 1_000_000_000)
        return {
            "source": {
                "name": self.workbook_path.name,
                "path": str(self.workbook_path),
                "modified_iso": modified.isoformat(timespec="seconds"),
                "modified_display": modified.strftime("%d %b %Y, %H:%M:%S"),
                "version": str(mtime_ns),
                "size": size,
            },
            "functions": functions,
            "function_order": ordered_function_names,
            "defaults": {"function": default_function, "year": default_year},
            "generated_iso": datetime.now().isoformat(timespec="seconds"),
        }

    def snapshot(self, force: bool = False) -> dict[str, Any]:
        with self._lock:
            current_mtime, current_size = self._version()
            changed = (
                force
                or self._snapshot is None
                or current_mtime != self._mtime_ns
                or current_size != self._size
            )
            if changed:
                try:
                    loaded = self._load()
                    self._snapshot = loaded
                    self._mtime_ns = current_mtime
                    self._size = current_size
                    self._last_error = None
                except Exception as exc:  # Keep the last valid dashboard available.
                    self._last_error = str(exc)
                    if self._snapshot is None:
                        raise
            assert self._snapshot is not None
            result = dict(self._snapshot)
            result["last_error"] = self._last_error
            return result

    def version_payload(self) -> dict[str, Any]:
        snapshot = self.snapshot()
        return {
            "version": snapshot["source"]["version"],
            "modified_iso": snapshot["source"]["modified_iso"],
            "modified_display": snapshot["source"]["modified_display"],
            "last_error": snapshot.get("last_error"),
        }

    def dashboard_payload(self, function_name: str | None = None, year: str | int | None = None) -> dict[str, Any]:
        snapshot = self.snapshot()
        functions = snapshot["functions"]
        if not functions:
            return {
                "source": snapshot["source"],
                "functions": [],
                "years": [],
                "selection": {"function": "", "year": ""},
                "summary": {},
                "themes": [],
                "last_error": snapshot.get("last_error"),
            }

        is_overall = (function_name is None or function_name == "_all_")
        if is_overall:
            selected_function = "_all_"
            years = {"2026": {}, "2027": {}, "2028": {}}
            requested_year = str(year) if year is not None else ""
            selected_year = requested_year if requested_year in years else "2026"
            
            all_themes = []
            for name in snapshot["function_order"]:
                func_data = functions[name]
                if selected_year in func_data["years"]:
                    all_themes.extend(func_data["years"][selected_year]["themes"])
            
            composite_summary = {
                "total": len(all_themes),
            }
            year_data = {"summary": composite_summary, "themes": all_themes}
        else:
            selected_function = function_name if function_name in functions else snapshot["defaults"]["function"]
            function_data = functions[selected_function]
            years = function_data["years"]
            requested_year = str(year) if year is not None else ""
            selected_year = requested_year if requested_year in years else (
                snapshot["defaults"]["year"] if selected_function == snapshot["defaults"]["function"] and snapshot["defaults"]["year"] in years else next(iter(years))
            )
            year_data = years[selected_year]

        function_options = []
        for name in snapshot["function_order"]:
            func_data = functions[name]
            year_data_func = func_data["years"].get(selected_year, {})
            themes_func = year_data_func.get("themes", [])
            
            total_actions_count = 0
            completed_count = 0
            active_count = 0
            risk_count = 0
            
            for theme in themes_func:
                for action in theme.get("actions", []):
                    total_actions_count += 1
                    if action["status"] == "complete":
                        completed_count += 1
                    elif action["status"] in {"not_started", "no_update"}:
                        pass
                    else:
                        active_count += 1
                        if action["status"] in {"at_risk", "blocked"}:
                            risk_count += 1
            
            key_themes = [
                {
                    "id": theme.get("id"),
                    "name": theme.get("name", ""),
                    "progress": theme.get("progress"),
                    "status": theme.get("status", "no_update"),
                    "status_label": theme.get("status_label", "No update"),
                    "actions_count": len(theme.get("actions", [])),
                }
                for theme in themes_func
            ]

            function_options.append({
                "name": name,
                "sheet_name": func_data["sheet_name"],
                "years": [int(key) for key in func_data["years"].keys()],
                "theme_count": len(themes_func),
                "total_actions_count": total_actions_count,
                "completed_count": completed_count,
                "active_count": active_count,
                "risk_count": risk_count,
                "key_themes": key_themes,
            })

        return {
            "source": snapshot["source"],
            "generated_iso": snapshot["generated_iso"],
            "functions": function_options,
            "years": [int(key) for key in years.keys()],
            "selection": {"function": selected_function, "year": int(selected_year)},
            "summary": year_data["summary"],
            "themes": year_data["themes"],
            "last_error": snapshot.get("last_error"),
        }
